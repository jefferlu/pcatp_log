"""
Page 6 — Fail Distribution
============================
Select one or more sessions to view each session's fail parameter
distribution across loops, and export the data to Excel.
"""
from __future__ import annotations

import io

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import PatternFill, Font

if not st.session_state.get("_username"):
    st.stop()

from components.sidebar import render_sidebar
from db.database import list_sessions, load_fail_values, load_all_results, load_device_info, load_fail_log_entries

render_sidebar(show_loop_selector=False, show_session_selector=False)

# ---------------------------------------------------------------------------
# Session selector
# ---------------------------------------------------------------------------
_username = st.session_state.get("_username", "")
_is_admin = st.session_state.get("_is_admin", False)
all_sessions = list_sessions(_username, is_admin=_is_admin)

if not all_sessions:
    st.info("No sessions imported yet.")
    st.stop()

_available_types = sorted({s["log_type"] for s in all_sessions if s.get("log_type")})
if len(_available_types) >= 2:
    _type_filter = st.radio(
        "Filter by type", _available_types,
        horizontal=True, key="cmp_type_filter",
    )
    _filtered = [s for s in all_sessions if s.get("log_type") == _type_filter]
else:
    _filtered = all_sessions

session_options = [s["session_id"] for s in _filtered]
selected_sessions = st.multiselect(
    "Select session(s)",
    options=session_options,
    default=[s for s in st.session_state.get("_cmp_sessions", []) if s in session_options],
    key="cmp_session_select",
)
st.session_state["_cmp_sessions"] = selected_sessions

if not selected_sessions:
    st.info("Select at least one session.")
    st.stop()

# ---------------------------------------------------------------------------
# Load fail values
# ---------------------------------------------------------------------------
with st.spinner("Loading fail data…"):
    fail_df = load_fail_values(selected_sessions)

if fail_df.empty:
    st.success("No FAIL records with numeric values found.")

st.divider()

# ---------------------------------------------------------------------------
# Per-session sections
# ---------------------------------------------------------------------------
for session_id in selected_sessions:
    sess_df = fail_df[fail_df["session_id"] == session_id]
    if sess_df.empty:
        st.warning(f"**{session_id}** — no numeric fail values.")
        continue

    params_sorted = sorted(sess_df["param"].unique())

    with st.expander(f"**{session_id}** — {len(params_sorted)} fail parameter(s)", expanded=True):
        # Summary table
        summary_rows = []
        for param in params_sorted:
            pdata = sess_df[sess_df["param"] == param]
            vals = pdata["numeric_value"]
            loop_count = pdata["loop_num"].nunique()
            lmin = pdata["limit_min"].dropna().iloc[0] if pdata["limit_min"].notna().any() else None
            lmax = pdata["limit_max"].dropna().iloc[0] if pdata["limit_max"].notna().any() else None
            test_name = pdata["test_name"].iloc[0] if not pdata.empty else ""
            sub_item  = pdata["sub_item"].iloc[0]  if not pdata.empty else ""
            summary_rows.append({
                "Test Name":        test_name,
                "Sub Item":         sub_item,
                "Total Fail Loops": int(loop_count),
                "Limit Min":        round(float(lmin), 4) if lmin is not None else None,
                "Limit Max":        round(float(lmax), 4) if lmax is not None else None,
                "Val Min":          round(float(vals.min()), 4),
                "Val Max":          round(float(vals.max()), 4),
                "Median":           round(float(np.median(vals)), 4),
                "Range":            round(float(vals.max() - vals.min()), 4),
            })
        summary_df = pd.DataFrame(summary_rows)

        st.dataframe(
            summary_df,
            hide_index=True,
            width="stretch",
            column_config={
                "Test Name":        st.column_config.TextColumn("Test Name",          width=180),
                "Sub Item":         st.column_config.TextColumn("Sub Item",           width=200),
                "Total Fail Loops": st.column_config.NumberColumn("Total Fail Loops", width=100),
                "Limit Min":        st.column_config.NumberColumn("Limit Min",        width=100, format="%.4f"),
                "Limit Max":        st.column_config.NumberColumn("Limit Max",        width=100, format="%.4f"),
                "Val Min":          st.column_config.NumberColumn("Val Min",          width=100, format="%.4f"),
                "Val Max":          st.column_config.NumberColumn("Val Max",          width=100, format="%.4f"),
                "Median":           st.column_config.NumberColumn("Median",           width=100, format="%.4f"),
                "Range":            st.column_config.NumberColumn("Range",            width=100, format="%.4f"),
            },
        )

        # Box plot
        plottable = [p for p in params_sorted if len(sess_df[sess_df["param"] == p]) >= 2]
        if plottable:
            fig = go.Figure()
            for param in plottable:
                pdata = sess_df[sess_df["param"] == param]
                fig.add_trace(go.Box(
                    y=pdata["numeric_value"].tolist(),
                    name=param,
                    text=[f"Loop {l}" for l in pdata["loop_num"].tolist()],
                    boxpoints="all",
                    jitter=0.3,
                    pointpos=0,
                    marker=dict(size=6, opacity=0.7),
                    hovertemplate="%{text}<br>Value: %{y}<extra></extra>",
                ))
            fig.update_layout(
                height=max(380, 55 * len(plottable)),
                margin=dict(t=20, b=120, l=60, r=20),
                xaxis=dict(tickangle=-35),
                yaxis=dict(title="Value"),
                showlegend=False,
            )
            with st.container(border=True):
                st.plotly_chart(fig, width="stretch", key=f"boxplot_{session_id}")

st.divider()

# ---------------------------------------------------------------------------
# Excel export — one sheet per session
# ---------------------------------------------------------------------------
st.subheader("Export to Excel")

_total_loops_map = {s["session_id"]: s["total_loops"] for s in all_sessions}

with st.spinner("Loading all results for export…"):
    all_results_df = load_all_results(selected_sessions)

with st.spinner("Loading log fail entries for export…"):
    log_fail_df = load_fail_log_entries(selected_sessions)

def _build_excel(
    sessions: list[str],
    fail_df: pd.DataFrame,
    all_df: pd.DataFrame,
    log_fail_df: pd.DataFrame,
) -> bytes:
    buf = io.BytesIO()
    _sum_fill  = PatternFill("solid", fgColor="4472C4")   # blue — summary header
    _raw_fill  = PatternFill("solid", fgColor="70AD47")   # green — raw header
    _log_fill  = PatternFill("solid", fgColor="FFC000")   # orange — log fail header
    _fail_fill = PatternFill("solid", fgColor="FFCCCC")   # red — fail rows
    _hdr_font  = Font(color="FFFFFF", bold=True)

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for session_id in sessions:
            sess_fail    = fail_df[fail_df["session_id"] == session_id]
            sess_all     = all_df[all_df["session_id"] == session_id]
            sess_log_fail = log_fail_df[log_fail_df["session_id"] == session_id] if not log_fail_df.empty else pd.DataFrame()
            if sess_all.empty and sess_log_fail.empty:
                continue

            total_loops = _total_loops_map.get(session_id, "—")

            # Summary (from fail data)
            summary_rows = []
            for param in sorted(sess_fail["param"].unique()) if not sess_fail.empty else []:
                pdata = sess_fail[sess_fail["param"] == param]
                vals  = pdata["numeric_value"]
                lmin  = pdata["limit_min"].dropna().iloc[0] if pdata["limit_min"].notna().any() else None
                lmax  = pdata["limit_max"].dropna().iloc[0] if pdata["limit_max"].notna().any() else None
                test_name = pdata["test_name"].iloc[0] if not pdata.empty else ""
                sub_item  = pdata["sub_item"].iloc[0]  if not pdata.empty else ""
                test_mode = pdata["test_mode"].iloc[0] if "test_mode" in pdata.columns and not pdata.empty else ""
                category  = pdata["category"].iloc[0]  if "category"  in pdata.columns and not pdata.empty else ""
                summary_rows.append({
                    "Test Name":        test_name,
                    "Sub Item":         sub_item,
                    "Test Mode":        test_mode,
                    "Category":         category,
                    "Total Fail Loops": int(vals.count()),
                    "Total Loops":      total_loops,
                    "Limit Min":        round(float(lmin), 4) if lmin is not None else "",
                    "Limit Max":        round(float(lmax), 4) if lmax is not None else "",
                    "Val Min":          round(float(vals.min()), 4),
                    "Val Max":          round(float(vals.max()), 4),
                    "Median":           round(float(np.median(vals)), 4),
                    "Range":            round(float(vals.max() - vals.min()), 4),
                })
            summary_df = pd.DataFrame(summary_rows)

            # Raw — all parameters (results + legacy_results), sorted by loop then test_name
            raw_df = pd.DataFrame()
            if not sess_all.empty:
                raw_df = sess_all[["loop_num", "test_mode", "category", "test_name", "sub_item", "result", "value"]].copy()
                raw_df.columns = ["Loop", "Test Mode", "Category", "Test Name", "Sub Item", "Result", "Value"]
                raw_df = raw_df.sort_values(["Loop", "Test Name", "Sub Item"]).reset_index(drop=True)

            # Log FAIL entries (from TXT log_entries, level='fail')
            log_df = pd.DataFrame()
            if not sess_log_fail.empty:
                log_df = sess_log_fail[["loop_num", "time_str", "module", "message"]].copy()
                log_df.columns = ["Loop", "Time", "Module", "Message"]
                log_df = log_df.reset_index(drop=True)

            # Write sheets
            # Layout:
            #   rows 1..N   = device info (key | value), one per info field
            #   row  N+1    = blank
            #   row  N+2    = "Fail Summary" title
            #   row  N+3    = summary header (blue)
            #   rows N+4..  = summary data
            #   (gap of 2)
            #   row  M      = "All Results" title
            #   row  M+1    = raw header (green)
            #   rows M+2..  = raw data
            #   (gap of 2)
            #   row  P      = "Log FAIL Entries" title
            #   row  P+1    = log header (orange)
            #   rows P+2..  = log data
            sheet_name  = session_id[:31]

            dev_info   = load_device_info(session_id)
            info_keys  = ["Project Name", "SW Version", "TB Version", "ZCU Version"]
            info_rows  = len(info_keys)            # always 4 rows reserved (blank if missing)
            info_offset = info_rows + 1            # blank row after info block

            sum_title_row = info_offset + 1        # 1-based
            sum_hdr_row   = sum_title_row + 1
            # startrow for to_excel is 0-based; summary header lands on sum_hdr_row
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False,
                                    startrow=sum_hdr_row - 1)
            else:
                # Write an empty placeholder so the sheet is created
                pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False,
                                        startrow=sum_hdr_row - 1)

            raw_title_row = sum_hdr_row + max(len(summary_df), 1) + 2
            raw_hdr_row   = raw_title_row + 1
            if not raw_df.empty:
                raw_df.to_excel(writer, sheet_name=sheet_name, index=False,
                                startrow=raw_hdr_row - 1)

            log_title_row = raw_hdr_row + max(len(raw_df), 1) + 2
            log_hdr_row   = log_title_row + 1
            if not log_df.empty:
                log_df.to_excel(writer, sheet_name=sheet_name, index=False,
                                startrow=log_hdr_row - 1)

            ws = writer.sheets[sheet_name]

            # Device info block
            for i, key in enumerate(info_keys):
                r = i + 1
                ws.cell(row=r, column=1).value = key
                ws.cell(row=r, column=2).value = dev_info.get(key, "")

            _title_font = Font(bold=True)

            # Fail Summary title
            ws.cell(row=sum_title_row, column=1).value = "Fail Summary"
            ws.cell(row=sum_title_row, column=1).font  = _title_font

            # All Results title
            ws.cell(row=raw_title_row, column=1).value = "All Results"
            ws.cell(row=raw_title_row, column=1).font  = _title_font

            # Log FAIL Entries title
            ws.cell(row=log_title_row, column=1).value = "Log FAIL Entries"
            ws.cell(row=log_title_row, column=1).font  = _title_font

            # Header colours
            for cell in ws[sum_hdr_row]:
                if cell.value is not None:
                    cell.fill = _sum_fill
                    cell.font = _hdr_font
            for cell in ws[raw_hdr_row]:
                if cell.value is not None:
                    cell.fill = _raw_fill
                    cell.font = _hdr_font
            for cell in ws[log_hdr_row]:
                if cell.value is not None:
                    cell.fill = _log_fill
                    cell.font = _hdr_font

            # Highlight FAIL rows in raw table
            if not raw_df.empty:
                for row_offset, result_val in enumerate(raw_df["Result"]):
                    if str(result_val).upper() == "FAIL":
                        excel_row = raw_hdr_row + 1 + row_offset
                        for cell in ws[excel_row]:
                            if cell.column <= len(raw_df.columns):
                                cell.fill = _fail_fill

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max(
                    (len(str(cell.value)) for cell in col if cell.value is not None),
                    default=8,
                )
                ws.column_dimensions[col[0].column_letter].width = max_len + 2

    buf.seek(0)
    return buf.read()

fname = "Fail_Distribution.xlsx"
with st.spinner("Building Excel…"):
    excel_bytes = _build_excel(selected_sessions, fail_df, all_results_df, log_fail_df)
st.download_button(
    label="Download Excel",
    data=excel_bytes,
    file_name=fname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary",
)
